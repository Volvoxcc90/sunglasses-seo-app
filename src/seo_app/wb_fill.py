import re
import time
import json
import random
from pathlib import Path
from typing import Callable, Optional, Tuple, List, Dict, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


TITLE_MAX_LEN = 60
DESC_MAX_LEN = 2000

DESC_LENGTH_RANGES = {
    "short":  (550, 850),
    "medium": (900, 1400),
    "long":   (1500, 2000),
}

SEO_LEVEL_COUNTS = {
    "soft":   {"core": 1, "tail": 1, "feature": 0, "extra": 0},
    "normal": {"core": 2, "tail": 1, "feature": 1, "extra": 0},
    "hard":   {"core": 2, "tail": 2, "feature": 1, "extra": 1},
}

FORBIDDEN_LABELS = [
    "сценарии:", "ключевые слова:", "форма:", "линза:", "коллекция:"
]

SLOGANS = [
    "Красивые", "Крутые", "Стильные", "Модные", "Молодёжные", "Трендовые",
    "Эффектные", "Дизайнерские", "Лаконичные", "Яркие", "Премиальные",
    "Удобные", "Лёгкие", "Универсальные", "Городские", "Летние",
    "Актуальные", "Смелые", "Элегантные", "Минималистичные",
    "Современные", "Ультрамодные", "Хитовые", "Культовые", "Фирменные",
    "Топовые", "Сочные", "Кайфовые", "Чёткие"
]

SEO_CORE = [
    "солнцезащитные очки", "солнечные очки", "очки солнцезащитные",
    "брендовые очки", "модные очки"
]

SEO_TAIL = [
    "очки для города", "очки для отпуска", "очки для вождения",
    "очки для путешествий", "аксессуар на лето", "очки унисекс",
    "инста очки", "очки из tiktok"
]

SEO_FEATURES = [
    "UV400", "поляризационные очки", "фотохромные очки",
    "зеркальные линзы", "градиентные линзы"
]

SEMANTIC_MATRIX = [
    {"focus": "Город/повседневка",   "must_tail": ["очки для города"],                          "add": ["брендовые очки"]},
    {"focus": "Вождение",            "must_tail": ["очки для вождения"],                        "add": ["очки солнцезащитные"]},
    {"focus": "Отпуск/путешествия",  "must_tail": ["очки для отпуска", "очки для путешествий"], "add": ["аксессуар на лето"]},
    {"focus": "Стиль/соцсети",       "must_tail": ["инста очки", "очки из tiktok"],             "add": ["модные очки"]},
    {"focus": "Универсальность",     "must_tail": ["очки унисекс"],                             "add": ["солнечные очки"]},
    {"focus": "Охват/ядро",          "must_tail": ["аксессуар на лето"],                        "add": ["очки солнцезащитные"]},
]

WB_SAFE_REPLACEMENTS = [
    (r"\bреплика\b", "стилизация"),
    (r"\bреплики\b", "стилизации"),
    (r"\bкопия\b", "дизайн в стиле"),
    (r"\bкопии\b", "дизайн в стиле"),
    (r"\b1\s*[:xх]\s*1\b", "в стиле оригинала"),
    (r"\bлюкс\b", "премиальный стиль"),
    (r"\blux\b", "премиальный стиль"),
    (r"\breplica\b", "style"),
]

STRICT_REWRITE = [
    (r"\b100%\b", "высокая"),
    (r"\bгарант(ия|ируем|ирует|ировано)\b", "обычно обеспечивает"),
    (r"\bлучши(й|е|ая|ие)\b", "отличный"),
    (r"\bидеальн(ый|ая|ое|ые)\b", "удачный"),
    (r"\bбезупречн(ый|ая|ое|ые)\b", "аккуратный"),
    (r"\bабсолютн(о|ый|ая|ое|ые)\b", "очень"),
    (r"\bоригинал(ьные|ьный|ьная|ьное)?\b", "фирменные"),
    (r"\bкак оригинал\b", "в стиле"),
    (r"\bлеч(ит|ат|ение)\b", "помогает чувствовать себя комфортнее"),
    (r"\bулучш(ает|ить)\s+зрение\b", "делает картинку более комфортной"),
    (r"\bснима(ет|ть)\s+усталост(ь|и)\b", "может снижать дискомфорт"),
    (r"\bзащищает\s+на\s+100%\b", "помогает защищать"),
    (r"\bтоп\s*1\b", "популярный выбор"),
    (r"\bномер\s*1\b", "популярный выбор"),
]
STRICT_DROP_PATTERNS = [
    r"\bподлинн(ый|ая|ое|ые)\b",
    r"\bсертифицир(ован|ованн|овано)\w*\b",
    r"\bофициальн(ый|ая|ое|ые)\b",
    r"\bгарантированн\w*\b",
]


# ==========================
# Бренды: ВАЖНО
# 1) если бренд уже кириллицей — оставляем
# 2) если латиницей — переводим ТОЛЬКО если есть в словаре
# 3) если нет в словаре — оставляем латиницу (не портим)
# ==========================
BRAND_RU_OVERRIDES = {
    "gucci": "Гуччи",
    "dior": "Диор",
    "prada": "Прада",
    "ray-ban": "Рэй-Бэн",
    "ray ban": "Рэй-Бэн",
    "cazal": "Казал",
    "versace": "Версаче",
    "chanel": "Шанель",
    "cartier": "Картье",
    "oakley": "Окли",
    "burberry": "Бёрберри",
    "balenciaga": "Баленсиага",
    "fendi": "Фенди",
    "givenchy": "Живанши",
    "saint laurent": "Сен-Лоран",
    "yves saint laurent": "Сен-Лоран",
    "dolce gabbana": "Дольче Габбана",
    "dolce & gabbana": "Дольче Габбана",
    "tom ford": "Том Форд",
    "gentle monster": "Джентл Монстер",
    "polaroid": "Полароид",
    "hugoboss": "Хьюго Босс",
    "hugo boss": "Хьюго Босс",
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _cut_no_word_break(text: str, max_len: int) -> str:
    text = _norm(text)
    if len(text) <= max_len:
        return text
    cut = text[:max_len].rsplit(" ", 1)[0]
    return cut.strip() if cut else text[:max_len].strip()


def _contains_cyrillic(s: str) -> bool:
    return bool(re.search(r"[А-Яа-яЁё]", s or ""))


def brand_display_name(brand: str) -> str:
    """Возвращает бренд для названия: кириллица если известный, иначе как ввели."""
    brand = _norm(brand)
    if not brand:
        return ""
    if _contains_cyrillic(brand):
        return brand

    key = brand.lower().replace("&", " ").replace("-", " ").strip()
    key = re.sub(r"\s+", " ", key)

    # сначала пробуем как есть, потом с дефисом/пробелом
    if key in BRAND_RU_OVERRIDES:
        return BRAND_RU_OVERRIDES[key]

    # ещё пробуем исходник без "лишнего"
    key2 = key.replace("  ", " ").strip()
    if key2 in BRAND_RU_OVERRIDES:
        return BRAND_RU_OVERRIDES[key2]

    # НЕ транслитерируем — оставляем латиницу, чтобы не было “кривого бренда”
    return brand


def _sun_term() -> str:
    return random.choice(["солнцезащитные очки", "солнечные очки"])


def _strip_forbidden(text: str) -> str:
    t = text
    for lab in FORBIDDEN_LABELS:
        t = re.sub(re.escape(lab), "", t, flags=re.IGNORECASE)
    return re.sub(r"\s{2,}", " ", t).strip()


def _apply_wb_safe(text: str) -> str:
    t = text
    for pattern, repl in WB_SAFE_REPLACEMENTS:
        t = re.sub(pattern, repl, t, flags=re.IGNORECASE)
    return re.sub(r"\s{2,}", " ", t).strip()


def _apply_wb_strict(text: str) -> str:
    t = text
    for pat in STRICT_DROP_PATTERNS:
        t = re.sub(pat, "", t, flags=re.IGNORECASE)
    for pat, repl in STRICT_REWRITE:
        t = re.sub(pat, repl, t, flags=re.IGNORECASE)
    t = re.sub(r"\s{2,}", " ", t).strip()
    t = re.sub(r"\s+,", ",", t)
    t = re.sub(r"\(\s*\)", "", t)
    return re.sub(r"\s{2,}", " ", t).strip()


def _first_n_words(text: str, n: int = 7) -> str:
    w = re.sub(r"[^0-9A-Za-zА-Яа-яёЁ ]+", " ", (text or "")).split()
    return " ".join(w[:n]).lower()


def _jaccard(a: str, b: str) -> float:
    def tok(x: str) -> set:
        x = re.sub(r"[^0-9A-Za-zА-Яа-яёЁ ]+", " ", x.lower())
        return {p for p in x.split() if len(p) > 2}
    A, B = tok(a), tok(b)
    return len(A & B) / max(1, len(A | B)) if A and B else 0.0


def _clamp_modes(style: str, seo_level: str, desc_length: str) -> Tuple[str, str, str]:
    style = (style or "neutral").lower().strip()
    if style not in {"neutral", "premium", "social"}:
        style = "neutral"

    seo_level = (seo_level or "normal").lower().strip()
    if seo_level not in {"soft", "normal", "hard"}:
        seo_level = "normal"

    desc_length = (desc_length or "medium").lower().strip()
    if desc_length not in {"short", "medium", "long"}:
        desc_length = "medium"

    return style, seo_level, desc_length


def build_titles_6(brand: str, shape: str, lens: str) -> List[str]:
    brand_show = brand_display_name(brand)
    shape = _norm(shape)
    lens = _norm(lens)

    # бренд рандомно: в 3 из 6 есть, в 3 из 6 нет
    flags = [True, True, True, False, False, False]
    random.shuffle(flags)

    templates = [
        "{slogan} {core} {brand}{shape}{lens}",
        "{slogan} {core} {shape}{brand}{lens}",
        "{slogan} {core} {lens}{brand}{shape}",
        "{slogan} {core} {brand}{lens}{shape}",
        "{slogan} {core} {shape}{lens}{brand}",
        "{slogan} {core} {lens}{shape}{brand}",
    ]

    local_slogans = random.sample(SLOGANS, k=6) if len(SLOGANS) >= 6 else [random.choice(SLOGANS) for _ in range(6)]
    used, out = set(), []

    for i in range(6):
        slogan = local_slogans[i]
        core = _sun_term()

        brand_part = (brand_show + " ") if (flags[i] and brand_show) else ""
        shape_part = (shape + " ") if (shape and random.random() < 0.55) else ""
        lens_part = (lens + " ") if (lens and random.random() < 0.70) else ""

        raw = templates[i].format(slogan=slogan, core=core, brand=brand_part, shape=shape_part, lens=lens_part)
        title = _cut_no_word_break(raw, TITLE_MAX_LEN)

        tries = 0
        while title in used and tries < 8:
            slogan = random.choice(SLOGANS)
            core = _sun_term()
            raw = templates[i].format(slogan=slogan, core=core, brand=brand_part, shape=shape_part, lens=lens_part)
            title = _cut_no_word_break(raw, TITLE_MAX_LEN)
            tries += 1

        used.add(title)
        out.append(title)

    return out


def pick_best_title(titles: List[str], last_slogan: str, recent_starts: List[str]) -> str:
    def slogan_of(t: str) -> str:
        return (t.split(" ", 1)[0] if t else "").strip()

    def start4(t: str) -> str:
        return " ".join((t or "").split()[:4]).lower()

    scored = []
    for t in titles:
        L = len(t)
        score = -abs(55 - L)
        if last_slogan and slogan_of(t).lower() == last_slogan.lower():
            score -= 6
        if recent_starts and start4(t) in recent_starts:
            score -= 5
        scored.append((score, t))

    scored.sort(key=lambda x: x[0], reverse=True)
    return scored[0][1] if scored else (titles[0] if titles else "")


def _lens_fact(lens: str) -> str:
    l = (lens or "").lower()
    if "uv400" in l:
        return random.choice([
            "UV400 часто выбирают для комфорта в солнечную погоду: меньше хочется щуриться, глаза устают меньше.",
            "Защита UV400 — удобный ориентир, когда нужно комфортно носить очки и в городе, и в поездках.",
        ])
    if "поляр" in l:
        return random.choice([
            "Поляризация помогает уменьшить блики от асфальта, воды и стекла — особенно заметно в дороге и на открытых пространствах.",
            "Поляризационный эффект делает картинку более читаемой при ярком свете и может снижать дискомфорт.",
        ])
    if "фото" in l or "хамеле" in l:
        return random.choice([
            "Фотохромный эффект удобен, когда освещение меняется: на улице темнее, в помещении спокойнее.",
            "Фотохромные линзы подходят тем, кто часто выходит из помещения на улицу и обратно.",
        ])
    return ""


# ==========================
# AUTO-пол
# ==========================
def infer_gender_mode(shape: str, lens: str) -> str:
    s = (shape or "").lower()
    l = (lens or "").lower()
    if "кошач" in s or "cat" in s:
        return "Жен"
    if "авиатор" in s or "pilot" in s:
        return "Унисекс"
    if "оверсайз" in s:
        return "Жен"
    if "спорт" in s or "sport" in s or "oakley" in l:
        return "Муж"
    return "Унисекс"


def gender_phrase(gender_mode: str) -> str:
    g = (gender_mode or "Auto").strip()
    if g == "Жен":
        return random.choice([
            "Подходит для женских образов — от повседневных до более выразительных.",
            "Акцентная модель для женского гардероба: смотрится современно и легко сочетается.",
        ])
    if g == "Муж":
        return random.choice([
            "Хороший вариант для мужского гардероба: уместно в городе и в поездках.",
            "Для мужских образов — практично, аккуратно и без лишней вычурности.",
        ])
    if g == "Унисекс":
        return random.choice([
            "Унисекс-посадка: легко вписывается в разные стили и сочетания.",
            "Унисекс-формат: подходит под разные образы и сценарии.",
        ])
    return ""


def _choose_keywords(lens: str, seo_level: str, slot: Dict[str, Any]) -> Dict[str, List[str]]:
    cfg = SEO_LEVEL_COUNTS[seo_level]

    core = random.sample(SEO_CORE, k=cfg["core"])
    for extra in slot.get("add", []):
        if extra not in core and len(core) < max(2, cfg["core"] + 1):
            core.append(extra)

    tail: List[str] = []
    for m in slot.get("must_tail", []):
        if m not in tail:
            tail.append(m)
        if len(tail) >= cfg["tail"]:
            break
    while len(tail) < cfg["tail"]:
        cand = random.choice(SEO_TAIL)
        if cand not in tail:
            tail.append(cand)

    features: List[str] = []
    if cfg["feature"] > 0:
        l = (lens or "").lower()
        if "uv400" in l:
            features.append("UV400")
        elif "поляр" in l:
            features.append("поляризационные очки")
        elif "фото" in l or "хамеле" in l:
            features.append("фотохромные очки")
        else:
            features.append(random.choice(SEO_FEATURES))

    extra2: List[str] = []
    if cfg["extra"] > 0:
        extra2.append(random.choice(["очки унисекс", "брендовые очки"]))

    if not any(("солнцезащитные" in x or "солнечные" in x) for x in core):
        core[0] = _sun_term()

    return {"core": core, "tail": tail, "features": features, "extra": extra2}


def _scenario_text_by_slot(slot: Dict[str, Any]) -> str:
    focus = (slot.get("focus", "") or "").lower()
    if "вожд" in focus:
        pool = ["вождение", "поездки", "город", "путешествия"]
    elif "отпуск" in focus or "путешеств" in focus:
        pool = ["отпуск", "пляж", "путешествия", "выходные"]
    elif "соц" in focus or "стиль" in focus:
        pool = ["город", "кафе и встречи", "выходные", "прогулки"]
    else:
        pool = ["город", "прогулки", "повседневные дела", "поездки"]
    return ", ".join(random.sample(pool, k=min(4, len(pool))))


def seo_card(text: str) -> Dict[str, Any]:
    low = (text or "").lower()
    found_core = [k for k in SEO_CORE if k in low]
    found_tail = [k for k in SEO_TAIL if k in low]
    found_feat = [k for k in SEO_FEATURES if k.lower() in low]
    if "uv400" in low and "UV400" not in found_feat:
        found_feat.append("UV400")

    score = 0
    score += min(4, len(found_core)) * 3
    score += min(3, len(found_tail)) * 2
    score += min(2, len(found_feat)) * 2

    label = "🟢 сильная" if score >= 12 else ("🟡 норм" if score >= 8 else "🔴 слабая")
    return {"score": score, "label": label, "found_core": found_core, "found_tail": found_tail, "found_features": found_feat}


BORING_STARTS = [
    "эта модель", "очки легко", "если хочется", "хороший вариант", "модель выглядит",
    "очки", "если в поиске"
]

def template_penalty(text: str) -> int:
    t = (text or "").lower().strip()
    penalty = 0
    for bs in BORING_STARTS:
        if t.startswith(bs):
            penalty += 3
            break
    if t.count("легко") >= 2:
        penalty += 1
    if t.count("актуально") >= 2:
        penalty += 1
    if t.count("удобно") >= 3:
        penalty += 1
    sents = [x.strip() for x in re.split(r"[.!?]+", t) if x.strip()]
    if len(sents) >= 4:
        short = sum(1 for s in sents[:4] if len(s) < 55)
        if short >= 3:
            penalty += 2
    return penalty


def generate_description_one(
    brand: str,
    shape: str,
    lens: str,
    collection: str,
    style: str,
    seo_level: str,
    desc_length: str,
    slot: Dict[str, Any],
    recent_desc_starts: List[str],
    wb_safe_mode: bool,
    wb_strict: bool,
    gender_mode: str,
) -> str:
    brand = _norm(brand)
    shape = _norm(shape)
    lens = _norm(lens)
    collection = _norm(collection)
    style, seo_level, desc_length = _clamp_modes(style, seo_level, desc_length)

    min_len, max_len = DESC_LENGTH_RANGES[desc_length]
    kw = _choose_keywords(lens, seo_level, slot)
    scen_txt = _scenario_text_by_slot(slot)

    gmode = (gender_mode or "Auto").strip()
    if gmode == "Auto":
        gmode = infer_gender_mode(shape, lens)

    g_text = ""
    if random.random() < 0.65:
        g_text = gender_phrase(gmode)

    openers = []
    if brand:
        openers += [
            f"{brand} — аксессуар, который делает образ собраннее и помогает чувствовать себя комфортно в солнечный день.",
            f"Очки {brand} добавляют уверенности: выглядят актуально и уместно, когда нужен летний акцент без перегруза.",
            f"Когда хочется подчеркнуть стиль — {brand} дают заметный эффект и при этом остаются удобными в повседневности.",
        ]
    else:
        openers += [
            "Аккуратный аксессуар на яркие дни: подчёркивает образ и даёт ощущение комфорта при солнечном свете.",
            "Удачная модель на каждый день: легко сочетается и выглядит свежо в городской среде.",
        ]
    opener = random.choice(openers)

    design = random.choice([
        f"Дизайн с {shape.lower()} линиями подчёркивает черты лица и делает образ более выразительным." if shape else
        "Дизайн подчёркивает черты лица и делает образ более выразительным.",
        "Оправа выглядит современно и хорошо сочетается с базовой одеждой и летними образами.",
        "Линии оправы смотрятся аккуратно и “собирают” образ даже без дополнительных аксессуаров."
    ])

    lenses_block = random.choice([
        f"Линзы {lens} дают комфорт при ярком солнце и подходят для активного дня." if lens else
        "Линзы дают комфорт при ярком солнце и подходят для активного дня.",
        f"С {lens} меньше хочется щуриться на улице, а дневной свет воспринимается спокойнее — особенно в городе и в дороге." if lens else
        "Дневной свет воспринимается спокойнее — особенно в городе и в дороге."
    ])

    fact = _lens_fact(lens)
    season = ""
    if collection and random.random() < 0.85:
        season = random.choice([
            f"Сезон {collection} — время лёгких деталей: модель выглядит свежо и уместно в повседневных образах.",
            f"Актуально на {collection}: можно носить каждый день и сохранять ощущение трендовой вещи.",
        ])

    if style == "premium":
        vibe = random.choice([
            "Визуально модель выглядит дороже за счёт чистых линий и аккуратных пропорций — образ получается уверенным.",
            "Сдержанный премиальный акцент: не спорит с другими деталями, но усиливает общий стиль.",
        ])
    elif style == "social":
        vibe = random.choice([
            "В кадре смотрится эффектно: добавляет летний вайб и делает образ более выразительным буквально за секунду.",
            "Хорошо “заходит” в фото: простой апгрейд, который сразу считывается как тренд.",
        ])
    else:
        vibe = random.choice([
            "Универсальный вариант на каждый день: легко сочетать с одеждой и не думать, подходит ли под образ.",
            "Практично и удобно: можно носить целый день и при этом выглядеть аккуратно и актуально.",
        ])

    core_str = ", ".join(kw["core"])
    tail_str = ", ".join(kw["tail"])
    feat_str = f" Часто такие модели ищут по запросу “{kw['features'][0]}”." if kw["features"] else ""
    extra_str = f" Также это может подойти как {kw['extra'][0]} — многое зависит от посадки." if kw["extra"] else ""

    tail = (
        f"Подходит для {scen_txt}. "
        f"Если в поиске нужны {core_str} и {tail_str}, здесь это совпадает с реальным удобством, а не только с картинкой."
        f"{feat_str}{extra_str}"
    )

    parts = [opener, design, lenses_block]
    if g_text:
        parts.append(g_text)

    if desc_length in {"medium", "long"}:
        if fact and random.random() < 0.9:
            parts.append(fact)
        if season and random.random() < 0.85:
            parts.append(season)
        parts.append(vibe)

    if desc_length == "long":
        parts.append(random.choice([
            "Носить удобно: аксессуар заметно “собирает” образ и помогает чувствовать себя увереннее на улице.",
            "Легко сочетаются с повседневной одеждой, когда хочется выглядеть современно без лишних усилий.",
            "Уместны и в городе, и в отпуске: добавляют уверенности и делают образ более цельным.",
        ]))

    parts.append(tail)

    mid = parts[1:-1]
    random.shuffle(mid)
    text = " ".join([parts[0]] + mid + [parts[-1]])
    text = _strip_forbidden(text)

    start = _first_n_words(text, 7)
    tries = 0
    while start in recent_desc_starts and tries < 6:
        parts[0] = random.choice(openers)
        mid = parts[1:-1]
        random.shuffle(mid)
        text = " ".join([parts[0]] + mid + [parts[-1]])
        text = _strip_forbidden(text)
        start = _first_n_words(text, 7)
        tries += 1

    if wb_safe_mode:
        text = _apply_wb_safe(text)
    if wb_strict:
        text = _apply_wb_strict(text)

    if len(text) > max_len:
        text = _cut_no_word_break(text, max_len)

    if len(text) < min_len and desc_length != "short":
        add = random.choice([
            "Это тот аксессуар, который легко носить каждый день и который заметно усиливает стиль.",
            "Модель выглядит уместно и не требует сложных сочетаний — надел и пошёл.",
            "Хороший баланс: и про внешний вид, и про комфорт, без лишней показухи.",
        ])
        text = _cut_no_word_break(text + " " + add, max_len)

    return _cut_no_word_break(text, min(DESC_MAX_LEN, max_len))


def pick_best_description(
    candidates: List[str],
    prev_desc: List[str],
    slot: Dict[str, Any],
    seo_level: str,
) -> Tuple[str, Dict[str, Any]]:
    best = candidates[0]
    best_meta: Dict[str, Any] = {}

    must_tail = slot.get("must_tail", [])
    must_tail_low = [m.lower() for m in must_tail]

    for text in candidates:
        sc = seo_card(text)
        penalty = template_penalty(text)

        low = text.lower()
        must_ok = sum(1 for m in must_tail_low if m in low)
        must_bonus = must_ok * 3

        if prev_desc:
            max_sim = max(_jaccard(text, p) for p in prev_desc)
        else:
            max_sim = 0.0
        uniq_bonus = int((1.0 - max_sim) * 6)

        seo_weight = 1.0 if seo_level != "hard" else 1.15
        score = int(sc["score"] * seo_weight) + must_bonus + uniq_bonus - penalty

        meta = {
            "score_total": score,
            "seo": sc,
            "template_penalty": penalty,
            "must_tail_hits": must_ok,
            "max_similarity": round(max_sim, 3),
        }

        if score > best_meta.get("score_total", -10**9):
            best = text
            best_meta = meta

    return best, best_meta


def find_header_row_and_cols(ws: Worksheet) -> Tuple[int, int, int]:
    for r in range(1, 16):
        name_col = desc_col = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                lv = v.lower()
                if "наименование" in lv:
                    name_col = c
                if "описание" in lv:
                    desc_col = c
        if name_col and desc_col:
            return r, name_col, desc_col
    raise ValueError("Не найдены колонки Наименование / Описание (первые 15 строк).")


def fill_wb_template(
    input_xlsx: str,
    brand: str,
    shape: str,
    lens_features: str,
    collection: str,
    style: str = "neutral",
    progress_callback: Optional[Callable[[int], None]] = None,
    seo_level: str = "normal",
    desc_length: str = "medium",
    wb_safe_mode: bool = True,
    wb_strict: bool = True,
    gender_mode: str = "Auto",
) -> Tuple[str, int, str]:
    random.seed(time.time())

    wb = load_workbook(input_xlsx)
    ws = wb.active

    header_row, col_name, col_desc = find_header_row_and_cols(ws)
    start = header_row + 1
    end = ws.max_row

    style, seo_level, desc_length = _clamp_modes(style, seo_level, desc_length)

    prev_desc: List[str] = []
    recent_desc_starts: List[str] = []
    recent_title_starts: List[str] = []
    last_title_slogan = ""

    total = max(1, end - start + 1)
    done = 0

    report: Dict[str, Any] = {
        "version": "v8-brand-fix",
        "input_file": str(input_xlsx),
        "settings": {
            "brand": brand,
            "shape": shape,
            "lens": lens_features,
            "collection": collection,
            "style": style,
            "seo_level": seo_level,
            "desc_length": desc_length,
            "wb_safe_mode": wb_safe_mode,
            "wb_strict": wb_strict,
            "gender_mode": gender_mode,
            "preview_candidates": 3
        },
        "rows": []
    }

    for idx, r in enumerate(range(start, end + 1)):
        slot = SEMANTIC_MATRIX[idx % len(SEMANTIC_MATRIX)]

        titles = build_titles_6(brand, shape, lens_features)
        title = pick_best_title(titles, last_title_slogan, recent_title_starts)

        last_title_slogan = (title.split(" ", 1)[0] if title else "")
        recent_title_starts.append(" ".join(title.split()[:4]).lower())
        if len(recent_title_starts) > 3:
            recent_title_starts.pop(0)

        candidates = [
            generate_description_one(
                brand=brand,
                shape=shape,
                lens=lens_features,
                collection=collection,
                style=style,
                seo_level=seo_level,
                desc_length=desc_length,
                slot=slot,
                recent_desc_starts=recent_desc_starts,
                wb_safe_mode=wb_safe_mode,
                wb_strict=wb_strict,
                gender_mode=gender_mode,
            )
            for _ in range(3)
        ]

        desc, pick_meta = pick_best_description(candidates, prev_desc, slot, seo_level)

        prev_desc.append(desc)
        recent_desc_starts.append(_first_n_words(desc, 7))
        if len(recent_desc_starts) > 6:
            recent_desc_starts.pop(0)

        ws.cell(r, col_name).value = title
        ws.cell(r, col_desc).value = desc

        report["rows"].append({
            "excel_row": r,
            "matrix_focus": slot.get("focus"),
            "title": title,
            "picked": pick_meta,
        })

        done += 1
        if progress_callback:
            progress_callback(int(done * 100 / total))

    out = Path(input_xlsx).with_name(Path(input_xlsx).stem + "_FILLED.xlsx")
    wb.save(out)

    report_json = out.with_suffix(".seo_report.json")
    report_txt = out.with_suffix(".seo_report.txt")
    report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    labels = [row["picked"]["seo"]["label"] for row in report["rows"]]
    green = labels.count("🟢 сильная")
    yellow = labels.count("🟡 норм")
    red = labels.count("🔴 слабая")

    lines = []
    lines.append("SEO REPORT")
    lines.append(f"Файл: {out.name}")
    lines.append(f"Safe: {'ON' if wb_safe_mode else 'OFF'} | Strict: {'ON' if wb_strict else 'OFF'} | Gender: {gender_mode}")
    lines.append(f"SEO: {seo_level} | Length: {desc_length} | Style: {style}")
    lines.append("")
    lines.append(f"Итог: 🟢 {green} | 🟡 {yellow} | 🔴 {red}")
    report_txt.write_text("\n".join(lines), encoding="utf-8")

    return str(out), done, str(report_json)
