# wb_fill.py (FULL REPLACE)
from __future__ import annotations

import json
import os
import random
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook


TITLE_MAX = 60
DESC_MAX = 2000

ProgressCB = Optional[Callable[[float], None]]

FORBIDDEN_LABELS_RE = re.compile(
    r"\b(Коллекция|Сценарии|Линзы|Линза|Форма|Ключевые\s*слова|Характеристики)\s*:\s*",
    flags=re.IGNORECASE
)

STOPWORDS_RU = {
    "и","в","во","на","а","но","что","это","как","для","по","из","к","с","со","при","от","до","у","же","не",
    "без","над","под","про","или","то","ли","же","та","те","этот","эта","эти","все","всё"
}


@dataclass
class FillParams:
    input_xlsx: str
    brand_lat: str
    shape: str
    lens: str
    collection: str
    holiday: str = ""

    seo_level: str = "high"               # low/normal/high
    gender_mode: str = "Auto"             # Auto/Женские/Мужские/Унисекс
    uniq_strength: int = 90               # 60..95
    brand_in_title_mode: str = "smart50"  # smart50/always/never

    max_fill_rows: int = 6
    skip_top_rows: int = 4

    output_index: int = 1
    output_total: int = 1
    between_files_slogan_lock: bool = True

    data_dir: str = ""


# ----------------------------
# Public API
# ----------------------------

def fill_wb_template(
    input_xlsx: str,
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    holiday: str = "",
    seo_level: str = "high",
    gender_mode: str = "Auto",
    uniq_strength: int = 90,
    brand_in_title_mode: str = "smart50",
    data_dir: str = "",
    max_fill_rows: int = 6,
    skip_top_rows: int = 4,
    output_index: int = 1,
    output_total: int = 1,
    between_files_slogan_lock: bool = True,
    progress_callback: ProgressCB = None,
) -> Tuple[str, int, dict]:
    # support call fill_wb_template(FillParams(...))
    if not isinstance(input_xlsx, str) and hasattr(input_xlsx, "__dict__"):
        return fill_wb_template(**dict(input_xlsx.__dict__))

    p = FillParams(
        input_xlsx=input_xlsx,
        brand_lat=brand_lat,
        shape=shape,
        lens=lens,
        collection=collection,
        holiday=holiday,
        seo_level=seo_level,
        gender_mode=gender_mode,
        uniq_strength=int(uniq_strength),
        brand_in_title_mode=brand_in_title_mode,
        data_dir=data_dir or "",
        max_fill_rows=int(max_fill_rows) if max_fill_rows else 6,
        skip_top_rows=int(skip_top_rows) if skip_top_rows else 4,
        output_index=int(output_index) if output_index else 1,
        output_total=int(output_total) if output_total else 1,
        between_files_slogan_lock=bool(between_files_slogan_lock),
    )
    return _fill_xlsx(p, progress_callback)


def generate_preview(
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    holiday: str = "",
    seo_level: str = "high",
    gender_mode: str = "Auto",
    uniq_strength: int = 90,
    brand_in_title_mode: str = "smart50",
    data_dir: str = "",
    count: int = 3,
) -> List[Tuple[str, str]]:
    st = _RunState(
        seed=_seed_for("preview"),
        data_dir=data_dir,
        between_files_slogan_lock=False,
    )
    items = []
    for _ in range(max(1, int(count))):
        title, desc = _generate_pair(
            brand_lat=brand_lat,
            shape=shape,
            lens=lens,
            collection=collection,
            holiday=holiday,
            seo_level=seo_level,
            gender_mode=gender_mode,
            uniq_strength=int(uniq_strength),
            brand_in_title_mode=brand_in_title_mode,
            st=st,
        )
        items.append((title, desc))
    return items


# ----------------------------
# Excel Fill
# ----------------------------

def _fill_xlsx(p: FillParams, progress_callback: ProgressCB) -> Tuple[str, int, dict]:
    in_path = Path(p.input_xlsx)
    if not in_path.exists():
        raise FileNotFoundError(str(in_path))

    wb = load_workbook(str(in_path))
    ws = wb.active

    name_col, desc_col = _find_cols(ws, header_rows=p.skip_top_rows)
    if not name_col or not desc_col:
        raise RuntimeError("не найдены колонки наименование и/или описание")

    rows_to_fill = 6  # строго 6
    start_row = p.skip_top_rows + 1
    end_row = start_row + rows_to_fill - 1

    out_path = _make_output_name(in_path, p.output_index, p.output_total)

    st = _RunState(
        seed=_seed_for(in_path.name + f"#{p.output_index}"),
        data_dir=p.data_dir,
        between_files_slogan_lock=p.between_files_slogan_lock,
    )

    descs = _generate_unique_descs(
        brand_lat=p.brand_lat,
        shape=p.shape,
        lens=p.lens,
        collection=p.collection,
        holiday=p.holiday,
        seo_level=p.seo_level,
        gender_mode=p.gender_mode,
        uniq_strength=p.uniq_strength,
        need=rows_to_fill,
        st=st,
    )

    done = 0
    for i, r in enumerate(range(start_row, end_row + 1)):
        title = _gen_title(
            brand_lat=p.brand_lat,
            shape=p.shape,
            lens=p.lens,
            brand_in_title_mode=p.brand_in_title_mode,
            st=st,
        )
        ws.cell(row=r, column=name_col).value = title
        ws.cell(row=r, column=desc_col).value = descs[i]

        done += 1
        if progress_callback:
            progress_callback((done / rows_to_fill) * 100.0)

    wb.save(str(out_path))
    return str(out_path), rows_to_fill, {"rows_filled": rows_to_fill, "output": str(out_path)}


def _find_cols(ws, header_rows: int) -> Tuple[Optional[int], Optional[int]]:
    want_name = {
        "наименование", "название", "наименование товара", "название товара", "title"
    }
    want_desc = {
        "описание", "описание товара", "description", "desc"
    }

    name_col = None
    desc_col = None

    for r in range(1, max(1, int(header_rows)) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            s = s.replace("ё", "е")
            s = re.sub(r"\s+", " ", s)
            if name_col is None and s in want_name:
                name_col = c
            if desc_col is None and s in want_desc:
                desc_col = c
        if name_col and desc_col:
            break

    return name_col, desc_col


def _make_output_name(in_path: Path, idx: int, total: int) -> Path:
    if total <= 1:
        return in_path.with_name(in_path.stem + "_filled.xlsx")
    width = max(2, len(str(total)))
    return in_path.with_name(in_path.stem + f"_filled_{idx:0{width}d}.xlsx")


# ----------------------------
# Random State + Locks
# ----------------------------

def _seed_for(seed_extra: str) -> int:
    mix = int.from_bytes(os.urandom(8), "big")
    h = (hash(seed_extra) & 0xFFFFFFFF)
    return (mix ^ h) & 0xFFFFFFFF


class _RunState:
    def __init__(self, seed: int, data_dir: str, between_files_slogan_lock: bool):
        self.rng = random.Random(seed)
        self.data_dir = data_dir or ""
        self.between_files_slogan_lock = between_files_slogan_lock

        self.used_title_sigs: Set[str] = set()
        self.used_desc_prefixes: Set[str] = set()
        self.used_desc_sigs: Set[str] = set()

        self.global_slogans: Set[str] = set()
        self.lock_path: Optional[Path] = None
        if between_files_slogan_lock and self.data_dir:
            self.lock_path = Path(self.data_dir) / "slogan_lock.json"
            self._load_lock()

    def _load_lock(self):
        if not self.lock_path or not self.lock_path.exists():
            return
        try:
            data = json.loads(self.lock_path.read_text(encoding="utf-8"))
            if isinstance(data, dict) and isinstance(data.get("slogans"), list):
                self.global_slogans = set(str(x).strip().lower() for x in data["slogans"])
        except Exception:
            pass

    def save_lock(self):
        if not self.lock_path:
            return
        try:
            payload = {"slogans": sorted(self.global_slogans)}
            self.lock_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass


# ----------------------------
# Brand RU map
# ----------------------------

def _nk(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("&", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _load_brand_ru_map(data_dir: str) -> Dict[str, str]:
    if not data_dir:
        return {}
    p = Path(data_dir) / "brands_ru.json"
    if not p.exists():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
        out = {}
        if isinstance(raw, dict):
            for k, v in raw.items():
                out[_nk(k)] = str(v).strip()
        return out
    except Exception:
        return {}


def _brand_ru(brand_lat: str, data_dir: str) -> str:
    mp = _load_brand_ru_map(data_dir)
    ru = mp.get(_nk(brand_lat))
    return (ru or brand_lat or "").strip()


# ----------------------------
# Text helpers
# ----------------------------

def _cap_first(text: str) -> str:
    t = (text or "").strip()
    t = re.sub(r"^[\s\-\–\—\•\·\.\,]+", "", t).strip()
    if not t:
        return ""
    return t[0].upper() + t[1:]


def _cut_no_break_words(text: str, limit: int) -> str:
    t = (text or "").strip()
    if len(t) <= limit:
        return t
    cut = t[:limit]
    if " " not in cut:
        return cut
    return cut.rsplit(" ", 1)[0].strip()


def _normalize_plain(text: str) -> str:
    t = (text or "").lower()
    t = FORBIDDEN_LABELS_RE.sub("", t)
    t = re.sub(r"[^a-zа-яё0-9\s\-]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _tokens(text: str) -> Set[str]:
    parts = _normalize_plain(text).split()
    return {w for w in parts if len(w) >= 3 and w not in STOPWORDS_RU}


def _jaccard(a: str, b: str) -> float:
    A = _tokens(a)
    B = _tokens(b)
    if not A or not B:
        return 0.0
    return len(A & B) / max(1, len(A | B))


def _desc_signature(text: str) -> str:
    t = _normalize_plain(text)
    words = t.split()
    pref = " ".join(words[:22])
    freq: Dict[str, int] = {}
    for w in words:
        if len(w) < 4 or w in STOPWORDS_RU:
            continue
        freq[w] = freq.get(w, 0) + 1
    top = " ".join([w for w, _ in sorted(freq.items(), key=lambda x: (-x[1], x[0]))[:6]])
    return (pref + " | " + top).strip()


def _uni_thr(uniq_strength: int) -> float:
    s = max(60, min(95, int(uniq_strength)))
    base = 0.78 - (s - 60) * (0.26 / 35.0)
    return max(0.38, base - 0.18)  # строго


# ----------------------------
# Title generator
# ----------------------------

def _gen_title(brand_lat: str, shape: str, lens: str, brand_in_title_mode: str, st: _RunState) -> str:
    slogans = [
        "Красивые","Крутые","Стильные","Модные","Молодёжные","Трендовые","Дизайнерские",
        "Эффектные","Лаконичные","Яркие","Удобные","Лёгкие","Актуальные","Премиальные",
        "Классные","Сочные","Смелые","Элегантные","Аккуратные","Статусные","Городские",
        "Летние","Повседневные","Универсальные","Топовые","Хитовые","С характером",
        "На каждый день","На лето","В тренде сезона"
    ]

    st.rng.shuffle(slogans)
    slogan = None
    for _ in range(80):
        cand = st.rng.choice(slogans).strip()
        key = cand.lower()
        if key in st.global_slogans:
            continue
        slogan = cand
        break
    if slogan is None:
        slogan = st.rng.choice(slogans).strip()

    if st.between_files_slogan_lock:
        st.global_slogans.add(slogan.lower())
        st.save_lock()

    sun = "солнцезащитные очки" if st.rng.random() < 0.6 else "солнечные очки"

    brand_ru = _brand_ru(brand_lat, st.data_dir)

    if brand_in_title_mode == "always":
        put_brand = True
    elif brand_in_title_mode == "never":
        put_brand = False
    else:
        put_brand = (st.rng.random() < 0.5)

    parts = [slogan, sun]
    if shape and st.rng.random() < 0.55:
        parts.append(shape)
    if put_brand and brand_ru:
        parts.append(brand_ru)
    if lens and st.rng.random() < 0.75:
        parts.append(lens)

    title = re.sub(r"\s+", " ", " ".join([p for p in parts if p]).strip())

    # anti repeat
    for _ in range(200):
        sig = _normalize_plain(title)
        if sig not in st.used_title_sigs:
            st.used_title_sigs.add(sig)
            break
        # small reshuffle
        st.rng.shuffle(parts[2:])
        title = re.sub(r"\s+", " ", " ".join([p for p in parts if p]).strip())

    while len(title) > TITLE_MAX:
        toks = title.split()
        if len(toks) <= 2:
            break
        toks.pop()
        title = " ".join(toks)

    return title


# ----------------------------
# Description generator (WB seller style + holiday injection)
# ----------------------------

def _build_desc_like_user(
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    holiday: str,
    seo_level: str,
    gender_mode: str,
    variant_id: int,
    st: _RunState,
) -> Tuple[str, str]:
    brand_lat = (brand_lat or "").strip()
    shape = (shape or "").strip()
    lens = (lens or "").strip()
    collection = (collection or "").strip()
    holiday = (holiday or "").strip()

    core1 = st.rng.choice(["солнцезащитные очки", "солнечные очки", "очки солнцезащитные"])
    core2 = st.rng.choice(["имиджевые очки", "модные очки", "брендовые очки", "трендовые очки"])

    if gender_mode == "Женские":
        gender_kw = "очки солнцезащитные женские"
    elif gender_mode == "Мужские":
        gender_kw = "очки солнцезащитные мужские"
    elif gender_mode == "Унисекс":
        gender_kw = "очки унисекс"
    else:
        gender_kw = st.rng.choice(["очки солнцезащитные женские", "очки солнцезащитные мужские", "очки унисекс"])

    seo_level = (seo_level or "high").lower().strip()
    seo_inserts = 1 if seo_level == "low" else 2 if seo_level == "normal" else 3

    starts = [
        f"{core2.capitalize()} {brand_lat} являются отличным дополнением к любому образу",
        f"Современные {core1} {brand_lat} сделают яркий акцент как в повседневном стиле, так и в нарядном",
        f"{core1.capitalize()} {brand_lat} помогают собрать образ и выглядеть стильно в солнечную погоду",
        f"Эти {core1} {brand_lat} легко сочетаются с одеждой и добавляют аккуратный стильный акцент",
        f"{core2.capitalize()} {brand_lat} — удачный вариант на каждый день и на сезон",
        f"{core1.capitalize()} {brand_lat} подойдут тем, кто любит стиль и комфорт без лишнего перегруза",
    ]

    b_style = [
        "Смотрятся современно и сразу обращают на себя внимание — вы будете притягивать взгляды окружающих",
        "Добавляют стильный акцент и делают образ более собранным",
        "Выглядят аккуратно и дорого, при этом легко сочетаются с одеждой",
        "Подходят и под базовый гардероб, и под более яркие сочетания",
        "Повседневный дизайн помогает выглядеть профессионально и стильно в течение дня",
    ]

    if shape:
        b_frame = [
            f"Красивая оправа {shape} подчёркивает черты лица и смотрится ровно",
            f"{shape.capitalize()} — удачная форма: подчёркивает стиль и не выглядит громоздко",
            f"Форма {shape} делает образ более выразительным и легко сочетается с одеждой",
            f"Оправа {shape} выглядит современно и обращает на себя внимание",
        ]
    else:
        b_frame = [
            "Оправа выглядит аккуратно и ровно сидит — носить комфортно в течение дня",
            "Дизайн оправы универсальный: легко вписывается в разные образы",
        ]

    if lens:
        b_lens = [
            f"Линзы {lens} дают комфорт при ярком солнце и подходят для активного дня",
            f"{lens} — хороший вариант для города и поездок, когда на улице ярко",
            f"С линзами {lens} проще в течение дня: солнце и отражения переносятся комфортнее",
        ]
    else:
        b_lens = [
            "Линзы комфортны в солнечную погоду — носить приятно в течение дня",
            "В яркий день глазам комфортнее — отличный вариант на повседневку",
        ]

    b_use = [
        "Подойдут для вождения, работы, учёбы, прогулок, отдыха, поездок и путешествий",
        "Можно носить в городе, в дороге, в отпуске, на пляже и на прогулках",
        "Удобны для повседневки: улица, дорога, отдых, прогулки, поездки",
        "Подойдут для работы за компьютером, прогулок, дороги и отдыха",
    ]

    b_unisex = [
        "Подойдут как для девушек, так и для мужчин — универсальный дизайн",
        "Модель унисекс: отлично дополняет и женский образ, и мужской",
        f"{gender_kw} — хороший выбор, если нужен универсальный аксессуар",
        "Универсальный вариант — подходит и девушкам, и мужчинам",
    ]

    b_gift = [
        "Отличный подарочный вариант для стильной девушки или парня",
        "Можно взять себе или на подарок — смотрятся презентабельно",
        "Хороший вариант в подарок: стильный аксессуар, который реально носят",
    ]

    b_coll = []
    if collection:
        b_coll = [
            f"На сезон {collection} модель выглядит актуально и легко вписывается в летний стиль",
            f"В сезоне {collection} такие очки особенно уместны: и в городе, и на отдыхе",
        ]

    b_note = [
        "Футляр может отличаться",
        "Комплектация может отличаться",
        "Оттенок может немного отличаться из-за настроек экрана",
    ]

    seo_phrases = [
        f"Такие {core1} часто выбирают как {core2}",
        f"{core1.capitalize()} удобны для города и отдыха",
        f"{core2.capitalize()} хорошо дополняют повседневный образ",
        f"{core1.capitalize()} подходят для отпуска и прогулок",
        f"{core2.capitalize()} — стильный акцент на каждый день",
    ]

    holiday_phrases = []
    if holiday:
        holiday_phrases = [
            f"Отлично подходит как подарок на {holiday} — аксессуар полезный и реально носится",
            f"Если ищешь подарок на {holiday}, это прям удачный вариант: стильно и практично",
            f"На {holiday} часто берут именно очки — вещь нужная на сезон и выглядит презентабельно",
        ]

    structs = [
        ("A", [st.rng.choice(starts), st.rng.choice(b_style), st.rng.choice(b_frame), st.rng.choice(b_lens)] + b_coll +
              [st.rng.choice(b_use), st.rng.choice(b_unisex), st.rng.choice(b_gift), st.rng.choice(b_note)]),
        ("B", [st.rng.choice(starts), st.rng.choice(b_frame), st.rng.choice(b_use), st.rng.choice(b_lens)] + b_coll +
              [st.rng.choice(b_style), st.rng.choice(b_unisex), st.rng.choice(b_gift), st.rng.choice(b_note)]),
        ("C", [st.rng.choice(starts), st.rng.choice(b_lens), st.rng.choice(b_style)] + b_coll +
              [st.rng.choice(b_frame), st.rng.choice(b_unisex), st.rng.choice(b_use), st.rng.choice(b_gift), st.rng.choice(b_note)]),
        ("D", [st.rng.choice(starts), st.rng.choice(b_unisex), st.rng.choice(b_frame), st.rng.choice(b_use)] + b_coll +
              [st.rng.choice(b_lens), st.rng.choice(b_style), st.rng.choice(b_gift), st.rng.choice(b_note)]),
        ("E", [st.rng.choice(starts), st.rng.choice(b_style), st.rng.choice(b_use), st.rng.choice(b_frame)] + b_coll +
              [st.rng.choice(b_lens), st.rng.choice(b_unisex), st.rng.choice(b_gift), st.rng.choice(b_note)]),
        ("F", [st.rng.choice(starts), st.rng.choice(b_frame), st.rng.choice(b_lens)] + b_coll +
              [st.rng.choice(b_use), st.rng.choice(b_style), st.rng.choice(b_unisex), st.rng.choice(b_gift), st.rng.choice(b_note)]),
    ]

    struct_key, parts = structs[variant_id % len(structs)]

    # SEO вставки внутрь текста
    for _ in range(seo_inserts):
        pos = st.rng.randint(1, max(1, len(parts) - 2))
        parts.insert(pos, st.rng.choice(seo_phrases))

    # Праздник — НЕ во все 6, а примерно в 2–3 из 6
    if holiday_phrases and st.rng.random() < 0.45:
        pos = st.rng.randint(3, max(3, len(parts) - 2))
        parts.insert(pos, st.rng.choice(holiday_phrases))

    text = ". ".join([p.strip().rstrip(".") for p in parts if p and p.strip()]).strip() + "."

    text = FORBIDDEN_LABELS_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = _cap_first(text)
    if len(text) > DESC_MAX:
        text = _cut_no_break_words(text, DESC_MAX)

    return text, struct_key


def _generate_unique_descs(
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    holiday: str,
    seo_level: str,
    gender_mode: str,
    uniq_strength: int,
    need: int,
    st: _RunState,
) -> List[str]:
    thr = _uni_thr(uniq_strength)
    out: List[str] = []
    used_structs: Set[str] = set()

    def prefix(text: str) -> str:
        t = _normalize_plain(text).split()
        return " ".join(t[:12])

    for i in range(need):
        best = None
        best_mx = 1.0
        best_struct = None

        for _ in range(320):
            cand, struct = _build_desc_like_user(
                brand_lat=brand_lat,
                shape=shape,
                lens=lens,
                collection=collection,
                holiday=holiday,
                seo_level=seo_level,
                gender_mode=gender_mode,
                variant_id=i,
                st=st,
            )

            pr = prefix(cand)
            if pr in st.used_desc_prefixes:
                continue

            sg = _desc_signature(cand)
            if sg in st.used_desc_sigs:
                continue

            if struct in used_structs and len(used_structs) < need:
                continue

            if not out:
                best = cand
                best_mx = 0.0
                best_struct = struct
                break

            mx = max(_jaccard(cand, prev) for prev in out)
            if mx <= thr:
                best = cand
                best_mx = mx
                best_struct = struct
                break

            if mx < best_mx:
                best = cand
                best_mx = mx
                best_struct = struct

        if best is None:
            best, best_struct = _build_desc_like_user(
                brand_lat, shape, lens, collection, holiday, seo_level, gender_mode, i, st
            )

        out.append(best)
        st.used_desc_prefixes.add(prefix(best))
        st.used_desc_sigs.add(_desc_signature(best))
        if best_struct:
            used_structs.add(best_struct)

    return out[:need]


def _generate_pair(
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    holiday: str,
    seo_level: str,
    gender_mode: str,
    uniq_strength: int,
    brand_in_title_mode: str,
    st: _RunState,
) -> Tuple[str, str]:
    title = _gen_title(brand_lat, shape, lens, brand_in_title_mode, st)
    desc, _ = _build_desc_like_user(
        brand_lat=brand_lat,
        shape=shape,
        lens=lens,
        collection=collection,
        holiday=holiday,
        seo_level=seo_level,
        gender_mode=gender_mode,
        variant_id=st.rng.randint(0, 999999),
        st=st,
    )
    return title, desc
